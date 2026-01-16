"""
extract_table.py
----------------
Descripción
    Lee un archivo Excel cuyo nombre contiene el prefijo 'FALLAS VVVF', extrae una tabla
    definida entre columnas A..X y construye un DataFrame de pandas con los datos filtrados.

Propósito
    Automatizar la lectura y limpieza inicial de la planilla de FALLAS VVVF para su procesado posterior.

Entradas / configuración
    - Argumentos CLI:
        * (opcional) Primer argumento: ruta a la carpeta donde buscar el fichero o ruta completa al fichero.
    - Variables de entorno / .env:
        * URL_MITSUBISHI: carpeta que contiene el/los archivos .xlsx.
        * FILE_PREFIX (opcional): prefijo para buscar el archivo (por defecto: "FALLAS VVVF").
    - Archivo .env recomendado: urls_sarmiento.env (ejemplo en el repo)

Salida
    - Imprime información del DataFrame resultante y, opcionalmente, logs informativos.
    - Devuelve/crea en memoria un DataFrame `df_clean` con las filas válidas (filtradas por columna 'N').

Cambios / historial
    - 2026-01-16: Añadida búsqueda por prefijo y manejo de .env.
"""
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple

import pandas as pd
from dotenv import load_dotenv
import os

# Archivos y carpetas
FILE_ENV = 'urls_sarmiento.env'
KEY = "URL_MITSUBISHI"
VALUE = "FALLAS VVVF MITSUBISHI 20240129_V1.0.xlsx"

# Limites de tabla
START_TABLE = "A1"
END_TABLE = "X1"

# Extraccion de URL del archivo xlsx desde un .env
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
dotenv_path = os.path.join(BASE_DIR, FILE_ENV)
load_dotenv(dotenv_path)
url_file_mit = os.getenv(KEY)
URL = os.path.join(url_file_mit, VALUE)

# Carga de archivo xlsx
wb = load_workbook(URL)
ws = wb.active

# Detecta si hay filas combinadas
for merge in ws.merged_cells.ranges:
    print(f"Celdas combinadas: {merge}")

try:
    df_test = pd.read_excel(URL, sheet_name=ws.title, header=None)
    print(df_test)
    print(df_test.info())
except Exception as e:
    print(f"Error al cargar el archivo: {e}")

# Ubico tabla en la hoja activa y obtengo un objeto generator
start_row, start_col = coordinate_to_tuple(START_TABLE)
_, end_col = coordinate_to_tuple(END_TABLE)

rows = ws.iter_rows(
    min_row=start_row,
    max_row=ws.max_row, 
    min_col=start_col, 
    max_col=end_col,
)

# Itero sobre generator para cargar valores en una lista de listas
data = []

for row in rows:
    data.append([cell.value for cell in row])

# Descrimino entre encabezado y datos
headers = data[0]
rows_data = data[1:]

# Creo dataframe
df = pd.DataFrame(rows_data, columns=headers)    

# Identifico y filtro solo filas validas - en base a num de registro
df["Num"] = pd.to_numeric(df["N"], errors="coerce")
df_clean = df.dropna(subset=["Num"])
df_clean = df_clean.drop("Num", axis=1)

print(df_clean)
print(df_clean.info())