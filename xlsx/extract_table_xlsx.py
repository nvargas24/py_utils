from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple

import pandas as pd

URL = 'E:\Documentos\Workspace\Data_libre\FALLAS VVVF MITSUBISHI 20240129_V1.0.xlsx'

# Carga de archivo xlsx
wb = load_workbook(URL)
ws = wb.active

# Detecta si hay filas combinadas
for merge in ws.merged_cells.ranges:
    print(f"Celdas combinadas: {merge}")

try:
    df_test = pd.read_excel(URL, sheet_name=ws.title, header=None)
    print(df_test)
    print(df_test.info())
except Exception as e:
    print(f"Error al cargar el archivo: {e}")


start_table = "A1"
end_table = "X1"

start_row, start_col = coordinate_to_tuple(start_table)
_, end_col = coordinate_to_tuple(end_table)

# Ubico tabla en la hoja activa y obtengo un objeto generator
rows = ws.iter_rows(
    min_row=start_row,
    max_row=ws.max_row, 
    min_col=start_col, 
    max_col=end_col,
)

# Itero sobre generator para cargar valores en una lista de listas
data = []

for row in rows:
    data.append([cell.value for cell in row])

# Descrimino entre encabezado y datos
headers = data[0]
rows_data = data[1:]

# Creo dataframe
df = pd.DataFrame(rows_data, columns=headers)    

# Identifico y filtro solo filas validas - en base a num de registro
df["Num"] = pd.to_numeric(df["N"], errors="coerce")
df_clean = df.dropna(subset=["Num"])
df_clean = df_clean.drop("Num", axis=1)

print(df_clean)
print(df_clean.info())